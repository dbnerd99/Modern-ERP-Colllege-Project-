import os
from flask import Flask, render_template, request, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

app = Flask(__name__)

@app.route('/') 
def home():
    return render_template('index.html')

@app.route('/calculate', methods=['GET', 'POST'])
def calculate():
    # Code for handling the upload request
    return render_template('calculate.html')

@app.route('/update_sheet', methods=['POST'])
def update_sheet():
    if request.method == 'POST':
        # Get the uploaded file
        file = request.files['file']

        # Load the workbook
        workbook = load_workbook(file)
        sheet = workbook.active
        sheet1 = workbook.active
        sheet= workbook.get_sheet_by_name('Sheet1')
        sheet1 = workbook.get_sheet_by_name('Sheet2')
        sheet2 = workbook.get_sheet_by_name('Sheet3')
        
    # Update the sheet with predefined formulas
    # Sheet1 CO-Attainment-1
    sheet['D80'] = 'target >50'
    sheet['K80'] = 'CO1'
    sheet['M80'] = 'CO2'
    sheet['O80'] = 'CO3'
    sheet['Q80'] = 'CO4'
    sheet['S80'] = 'CO5'
    sheet['U80'] = 'CO6'

    sheet['F86'] = 'CO1'
    sheet['G86'] = 'CO2'
    sheet['H86'] = 'CO3'
    sheet['I86'] = 'CO4'
    sheet['J86'] = 'CO5'
    sheet['K86'] = 'CO6'

    sheet['D86'] = 'Course Outcomes'
    sheet['D87'] = 'Internal Attainment'
    sheet['D88'] = 'External Attainment'
    sheet['D89'] = 'Internal Attainment with 20%'
    sheet['D90'] = 'External Attainment with 80%'
    sheet['D91'] = 'Overall Attainment 80:20'
    sheet['D92'] = 'Overall Attainment 80%'

    # Setting Font    
    font = Font(name='Calibri', size=11, bold=True, italic=False, color='000000')
    for row in sheet['D80:U80']:
        for cell in row:
            cell.font = font

    font = Font(name='Calibri', size=11, bold=True, italic=False, color='000000')
    for row in sheet['F86:K86']:
        for cell in row:
            cell.font = font

    font = Font(name='Calibri', size=11, bold=True, italic=True, color='000000')
    for row in sheet['D86:D92']:
        for cell in row:
            cell.font = font

    sheet["F80"] = '=COUNTIF(F13:F79, ">=50")'

    columns = ["K", "M", "O", "Q", "S", "U"]
    for column in columns:
        formula = f'=COUNTIF({column}13:{column}79, ">5")'
        cell = f'{column}81'
        sheet[cell] = formula

    columns = ["K", "M", "O", "Q", "S", "U"]
    for i, column in enumerate(columns):
        cell = chr(ord('F') + i) + '87'
        formula = f'={column}81*100/C9'
        sheet[cell] = formula

    columns = ["F", "G", "H", "I", "J", "K"]
    for column in columns:
        cell = f'{column}88'
        formula = f'=F80*100/F80'
        sheet[cell] = formula

    columns = ["F", "G", "H", "I", "J", "K"]
    for column in columns:
        cell = f'{column}89'
        formula = f'={column}87*20/100'
        sheet[cell] = formula

    columns = ["F", "G", "H", "I", "J", "K"]
    for column in columns:
        cell = f'{column}90'
        formula = f'={column}88*0.8'
        sheet[cell] = formula

    columns = ["F", "G", "H", "I", "J", "K"]
    for column in columns:
        cell = f'{column}91'
        formula = f'={column}89+{column}90'
        sheet[cell] = formula

    columns = ["F", "G", "H", "I", "J", "K"]
    for column in columns:
        cell = f'{column}92'
        formula = f'={column}91*0.8'
        sheet[cell] = formula

    # Setting Dimensions
    sheet.column_dimensions['D'].width = 20

    # Setting allignement
    for row in sheet['D80:U81']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adding Border
    def set_border(sheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    set_border(sheet, 'D80:U81') 

    def set_border(sheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    set_border(sheet, 'D86:K92')

    # Adding Colour
    _color1 = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    for row in sheet['D86:K92']:
        for cell in row:
            cell.fill = _color1

        _color2 = PatternFill(start_color="808000", end_color="808000", fill_type="solid")
    for row in sheet['D80:U81']:
        for cell in row:
            cell.fill = _color2

    # Sheet-2 Indirect Shift-1
    sheet1['A50'] = 'CO1'
    sheet1['A51'] = 'CO2'
    sheet1['A52'] = 'CO3'
    sheet1['A53'] = 'CO4'
    sheet1['A54'] = 'CO5'
    sheet1['A55'] = 'CO6'
    sheet1['A56'] = 'ALL'

    sheet1['B48'] = 5
    sheet1['C48'] = 4
    sheet1['D48'] = 3
    sheet1['E48'] = 2
    sheet1['F48'] = 1

    sheet1['B49'] = 'Excellent'
    sheet1['C49'] = 'Good'
    sheet1['D49'] = 'Average'
    sheet1['E49'] = 'Fair'
    sheet1['F49'] = 'Poor'
    sheet1['G49'] = 'Total Score'
    sheet1['H49'] = '% Score'
    sheet1['I49'] = 'Avg %'
    sheet1['J49'] = '20% Score'

    # Setting Allignment
    for row in sheet1['A48:J56']:
     for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adding color
     _color3 = PatternFill(start_color="993300", end_color="993300", fill_type="solid")
     for row in sheet1['A50:A56']:
        for cell in row:
         cell.fill = _color3

    _color4 = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    for row in sheet1['B48:F48']:
      for cell in row:
        cell.fill = _color4        

    columns = ["C", "D", "E", "F", "G", "H", "I"]
    start_row = 50
    for i, column in enumerate(columns):
        cell = f"B{start_row + i}"
        formula = f'=COUNTIF({column}9:{column}46, ">=5")'
        sheet1[cell] = formula

    columns = ["C", "D", "E", "F", "G", "H", "I"]
    start_row = 50
    for i, column in enumerate(columns):
        cell = f"C{start_row + i}"
        formula = f'=COUNTIF({column}9:{column}46, "=4")'
        sheet1[cell] = formula

    columns = ["C", "D", "E", "F", "G", "H", "I"]
    start_row = 50
    for i, column in enumerate(columns):
        cell = f"D{start_row + i}"
        formula = f'=COUNTIF({column}9:{column}46, "=3")'
        sheet1[cell] = formula

    columns = ["C", "D", "E", "F", "G", "H", "I"]
    start_row = 50
    for i, column in enumerate(columns):
        cell = f"E{start_row + i}"
        formula = f'=COUNTIF({column}9:{column}46, "=2")'
        sheet1[cell] = formula

    columns = ["C", "D", "E", "F", "G", "H", "I"]
    start_row = 50
    for i, column in enumerate(columns):
        cell = f"F{start_row + i}"
        formula = f'=COUNTIF({column}9:{column}46, "=1")'
        sheet1[cell] = formula

    for i in range(50, 57):
        formula = f"=B{i}*5+C{i}*4+D{i}*3+E{i}*2+F{i}*1"
        cell = f"G{i}"
        sheet1[cell] = formula

    for row in range(50, 57):
        sheet1[f'H{row}'] = f"=G{row}/(SUM(B{row}:F{row})*.05)"

    for i in range(50, 57):
        formula = f"=AVERAGE(H{i},H56)"
        cell = f"I{i}"
        sheet1[cell] = formula

    for row in range(50,57):
        sheet1[f'J{row}'] = f"=I{row}*20%"

    # Adding border
    def set_border(sheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in sheet1[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    set_border(sheet1, 'A48:J56')

    # Setting font
    font = Font(name='Calibri', size=11, bold=True, italic=True, color='000000')
    for row in sheet1['B49:J49']:
        for cell in row:
            cell.font = font

    font = Font(name='Calibri', size=11, bold=True, italic=True, color='000000')
    for row in sheet1['A50:A56']:
        for cell in row:
            cell.font = font


    cell_values = ['F', 'G', 'H', 'I', 'J', 'K']
    start_row = 67
    for i, cell_value in enumerate(cell_values):
        formula = f"='Sheet1'!{cell_value}88/100"
        cell = f"C{start_row + i}"
        sheet2[cell] = formula

    cell_values = ['F', 'G', 'H', 'I', 'J', 'K']
    start_row = 67
    for i, cell_value in enumerate(cell_values):
        formula = f"='Sheet1'!{cell_value}90/100"
        cell = f"E{start_row + i}"
        sheet2[cell] = formula

    cell_values = ['F', 'G', 'H', 'I', 'J', 'K']
    start_row = 67
    for i, cell_value in enumerate(cell_values):
        formula = f"='Sheet1'!{cell_value}87/100"
        cell = f"G{start_row + i}"
        sheet2[cell] = formula

    cell_values = ['F', 'G', 'H', 'I', 'J', 'K']
    start_row = 67
    for i, cell_value in enumerate(cell_values):
        formula = f"='Sheet1'!{cell_value}89/100"
        cell = f"I{start_row + i}"
        sheet2[cell] = formula

    cell_values = ['F', 'G', 'H', 'I', 'J', 'K']
    start_row = 67
    for i, cell_value in enumerate(cell_values):
        formula = f"='Sheet1'!{cell_value}91/100"
        cell = f"K{start_row + i}"
        sheet2[cell] = formula

    cell_values = ['F', 'G', 'H', 'I', 'J', 'K']
    start_row = 67
    for i, cell_value in enumerate(cell_values):
        formula = f"='Sheet1'!{cell_value}92/100"
        cell = f"M{start_row + i}"
        sheet2[cell] = formula

    start_row = 78
    start_index = 50
    for i in range(start_row, start_row + 6):
        formula = f"='Sheet2'!I{start_index}/100"
        cell = f"C{i}"
        sheet2[cell].value = formula
        start_index += 1

    start_row = 78
    start_index = 50
    for j in range(start_row, start_row + 6):
        formula = f"='Sheet2'!J{start_index}/100"
        cell = f"G{j}"
        sheet2[cell].value = formula
        start_index += 1

    start_column = 67
    for i in range(6):
        formula = f"=M{67 + i}"
        cell = chr(ord('D') + i) + '92'
        sheet2[cell] = formula

    start_column = 78
    for j in range(6):
        formula = f"=G{78 + j}"
        cell = chr(ord('D') + j) + '93'
        sheet2[cell] = formula

    columns_1 = ["D", "E", "F", "G", "H", "I"]
    for column in columns_1:
        cell = f'{column}94'
        formula = f'={column}92+{column}93'
        sheet2[cell] = formula

    start_row = 103
    start_column = 94
    for i in range(6):
        formula = f"={chr(ord('D') + i)}94"
        cell = f"B{start_row + i}"
        sheet2[cell] = formula

    columns = ['C', 'D', 'E','F','G','H','I','J','K','L','M','N']
    start_row = 103
    end_row = 108
    for column in columns:
        formula = f'=IF(COUNT({column}{start_row}:{column}{end_row})=0, "-", SUM({column}{start_row}:{column}{end_row})/COUNT({column}{start_row}:{column}{end_row}))'
        cell = f'{column}109'
        sheet2[cell] = formula

    columns = ['C', 'D', 'E','F','G','H','I','J','K','L','M','N']
    row = 103
    B103 = sheet2['B103'].value
    for column in columns:
        formula = f'=IF({column}{row}="-","-",{column}{row}*B103)'
        cell = f'{column}118'
        sheet2[cell] = formula

    row = 104
    B104 = sheet2['B104'].value
    for column in columns:
        formula = f'=IF({column}{row}="-","-",{column}{row}*B104)'
        cell = f'{column}119'
        sheet2[cell] = formula

    row = 105
    B105 = sheet2['B105'].value
    for column in columns:
        formula = f'=IF({column}{row}="-","-",{column}{row}*B105)'
        cell = f'{column}120'
        sheet2[cell] = formula

    row = 106
    B106 = sheet2['B106'].value
    for column in columns:
        formula = f'=IF({column}{row}="-","-",{column}{row}*B106)'
        cell = f'{column}121'
        sheet2[cell] = formula

    row = 107
    B107 = sheet2['B107'].value
    for column in columns:
        formula = f'=IF({column}{row}="-","-",{column}{row}*B107)'
        cell = f'{column}122'
        sheet2[cell] = formula

    columns = ['C', 'D', 'E','F','G','H','I','J','K','L','M','N']
    start_row = 118
    end_row = 122
    for column in columns:
        formula = f'=IF(COUNT({column}{start_row}:{column}{end_row})=0, "-", SUM({column}{start_row}:{column}{end_row})/COUNT({column}{start_row}:{column}{end_row}))'
        cell = f'{column}123'
        sheet2[cell] = formula

    columns_2 = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
    start_row = 123
    for column in columns_2:
        formula = f'={column}{start_row}'
        cell = f'{column}131'
        sheet2[cell] = formula
        
        # Save the updated workbook
    filename = 'updated_sheet.xlsx'
    workbook.save(filename)

        # Send the updated sheet as a downloadable file
    return send_from_directory(directory=os.getcwd(), path=filename, as_attachment=True)

    # Return an empty response if the method is not POST
    return ''

if __name__ == '__main__':
    app.run(port=5500)
